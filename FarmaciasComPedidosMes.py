import pandas as pd
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import shutil, os
from datetime import date
import datetime
import time
from collections import Counter

planilhaEditada = "KPI's Interno"
scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]

creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)

client = gspread.authorize(creds)

#Seleciona o arquivo que estarei visualizando / editando.
sheet2 = client.open(planilhaEditada).worksheet("KPI's")
sheet = client.open(planilhaEditada).worksheet("Farmácias") 

ano = '2021'
mes = '12'

#TO DO = RESOLVER COLUNAS
coluna = 13

opt = webdriver.ChromeOptions() 
opt.add_experimental_option("excludeSwitches", ["enable-automation"])
opt.add_experimental_option('useAutomationExtension', False)
opt.add_argument("--window-size=1034,620")

driver = webdriver.Chrome(options=opt, executable_path= r'C:\Users\Take4\Documents\Saves\chromedriver_win32\chromedriver.exe')

driver.get("https://homolog.farmazon.com.br")

mail_addressDjango = "-"
passwordDjango = "-"

#Autenticação no Django.
WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.NAME, "username"))).send_keys(mail_addressDjango)
WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.NAME, "password"))).send_keys(passwordDjango)
WebDriverWait(driver, 60).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div/form[2]/button'))).click()

driver.get('https://homolog.farmazon.com.br/api/orders/?format=json&limit=5000')

with open('farmacias.txt', 'w') as f:
    f.write(driver.page_source)

arquivo25 = (r'C:\Users\Take4\Desktop\KPI\\farmacias.txt')

#Arquivo TXT salvo é lido para ver quantas repeticoes do "create time" + ano + mes existem.
f25 = open(arquivo25)
file_contents25 = f25.read()

firstOrder = file_contents25.split('","create_time":"' + ano + "-" + mes, 1)[1]
lastOrder2 = firstOrder.split(',"url"', 1)[0]
lastOrder = lastOrder2.split(',{"id":', 1)[1]

ultimoPedidoDoMes = int(lastOrder) + 1


def lendoPedidos(pedidoMes):
    driver.get('https://homolog.farmazon.com.br/api/orders/'+ str(pedidoMes) +'/?format=json')
    with open('farmacias2.txt', 'w', encoding = 'utf-16') as f:
        f.write(driver.page_source)
    arquivos2 = (r'C:\Users\Take4\Desktop\KPI\\farmacias2.txt')
    f2 = open(arquivos2, encoding = 'utf-16')
    file_contents2 = f2.read()
    lendoData = file_contents2.count(',"create_time":"' + ano + '-' + mes)
    lendoData2 = file_contents2.count('},"create_time":"' + ano + '-' + mes)
    if lendoData == 1 or lendoData2 == 1:
        return ('true')
    else:
        return ('false')
    f2.close()


farmacias = []
while lendoPedidos(ultimoPedidoDoMes) == 'true':
    
    with open('recorrencias.txt', 'w', encoding = 'utf-16') as f:
        f.write(driver.page_source)
    arquivos3 = (r'C:\Users\Take4\Desktop\KPI\\recorrencias.txt')
    f3 = open(arquivos3, encoding = 'utf-16')
    file_contents3 = f3.read()
    customer1IDCount = file_contents3.count('"customer_id":"3"')
    customer2IDCount = file_contents3.count('"customer_id":"1023"')
    customer3IDCount = file_contents3.count('"customer_id":"1112"')
    customer4IDCount = file_contents3.count('"customer_id":"1"')
    customer5IDCount = file_contents3.count('"customer_id":"932"')
    customer6IDCount = file_contents3.count('"customer_id":"1167"')
    customer7IDCount = file_contents3.count('"customer_id":"5"')
    cancelCount = file_contents3.count('"cancelado"')
    realCount1 = file_contents3.count('},"create_time":"'+ano+'-'+mes)
    realCount2 = file_contents3.count('l,"create_time":"'+ano+'-'+mes)
    contagemEntregue = file_contents3.count('entregue')
    pagamentoContagem = file_contents3.count('"operator_payment_id":')
    storeNull = file_contents3.count(',"store":null,')
    if realCount1 == 1 or realCount2 == 1:
        if customer1IDCount == 1 or customer2IDCount == 1 or customer3IDCount == 1 or customer4IDCount == 1 or customer5IDCount == 1 or customer6IDCount == 1 or customer7IDCount == 1:
            pass
        elif cancelCount == 2:
            pass
        elif storeNull == 1:
            pass
        elif contagemEntregue == 2 or pagamentoContagem == 1:
            Id2 = file_contents3.split(':{"url":"', 1)[1]
            ID = Id2.split('","country":', 1)[0]
            ID3 = ID.split('","name":"', 1)[1]
            farmacias.append(ID3)
            print ("Pedido nº: " + str(ultimoPedidoDoMes) + " |||  Farmacia: " + str(ID3))
        else:
            #Azedou
            print ("Verificar: " + str(ultimoPedidoDoMes))
    else:

        f3.close()
        break
    
    ultimoPedidoDoMes -= 1
    
driver.close()
print (farmacias)

contagemRecorrenciasFarmacias = (Counter(farmacias))

print (contagemRecorrenciasFarmacias)

farmaciasComPeloMenos1Pedido = (len(contagemRecorrenciasFarmacias))

linhaFarmacias = 17

if ano == "2021":
    Mes = int(mes) + 19
elif ano == "2022":
    Mes = int(mes) + 31
    
sheet2.update_cell (linhaFarmacias, Mes, farmaciasComPeloMenos1Pedido)

######################################################################################################

def lendo (mes):
    mes = mes.replace('Counter', '')
    mes = mes.replace('({', '')
    mes = mes.replace('})', '')
    mes = mes.replace("'", "")
    mes = mes.replace(',', "','")
    mes = mes.replace("',' ", "','")
    mes = mes.replace("'", "")
    
    mes2 = mes.split(',')
    return mes2

Mes = str(contagemRecorrenciasFarmacias)



array = lendo(Mes)

index = 0
linha = len(sheet.get_all_values()) + 1


#WHILE LOOP PARA INSERIR NOVAS FARMACIAS NA PLANILHA CASO ELAS NAO ESTEJAM LA AINDA.
while index < len(array):
    farmacia = array[index].split(':', 1)[0]
    numerosFarmacia = array[index].split(': ', 1)[1]
   
    valor2 = sheet.cell(linha, 1).value

    print (farmacia)
    
    cell = sheet.findall(farmacia, in_column=1)
    
    cell = str(cell)

    if 'Cell' not in cell: 
        sheet.update_cell(linha, 1, farmacia)
        linha += 1
    
    else:
        pass

    index += 1
    
#WHILE LOOP PARA INSERIR QUANTIDADE DE PEDIDOS DE CADA FARMACIA.
linha2 = 2
index2 = 0

while index2 < len(array):
    farmacia = array[index2].split(':', 1)[0]
    numerosFarmacia = array[index2].split(': ', 1)[1]
   
    valor2 = sheet.cell(linha2, 1).value
    
    localizadorLinha = (sheet.find(farmacia))
    
    localizadorLinha1 = str(localizadorLinha)
    
    localizadorLinha2 = localizadorLinha1.replace('<Cell R', '').replace('C1', '').replace(farmacia, '').replace(" ''>", '')

    sheet.update_cell(localizadorLinha2, coluna, numerosFarmacia)

    index2 += 1
        
    
linha3 = 2

while linha3 < 19:
    valor3 = sheet.cell(linha3, coluna).value
    if len(valor3) == 0:
        sheet.update_cell(linha3, coluna, 0)
    linha3 += 1












