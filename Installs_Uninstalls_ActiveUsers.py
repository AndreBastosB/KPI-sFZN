import pandas as pd
from openpyxl import load_workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import shutil, os
from datetime import date
import datetime
import time


# AUTOMAÇÃO DOS SEGUINTES KPI'S NESTE CÓDIGO:
    # - NOVAS INSTALAÇÕES.
    # - USUÁRIOS ATIVOS.
    # - NOVOS CADASTROS
    # - DESINSTALAÇÕES.

#OBS: NECESSARIO RODAR CÓDIGO "DOWNLOAD REPORTS.PY" ANTES DESTE.


data = datetime.datetime.now()    

ano = str(data.year)

mes = str(data.month)

if (len(mes)) == 1:
    mes = '0' + mes
else:
    pass

diainicial = '01'

diafinal1 = (data.day)
diafinal2 = int(diafinal1) - 1
diafinal = str(diafinal2)

if (len(diafinal)) == 1:
    diafinal = '0' + diafinal
    

planilhaEditada = "KPI's Interno"
armazenamentoRelatorios = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+'\\'

if not os.path.exists(armazenamentoRelatorios):
    os.makedirs(armazenamentoRelatorios)
    

#AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET
#AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET - AUTENTICAÇÃO AO GOOGLE SHEET - 
scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]

creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)

client = gspread.authorize(creds)

#Seleciona o arquivo que estarei visualizando / editando.
sheet = client.open(planilhaEditada).sheet1

#FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE
#FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE - FORMATAÇÃO RELATORIOS APPLE

#Passo 1: Colocar as configurações do selenium webdriver da maneira correta a importar os relatorios.
opt = webdriver.ChromeOptions() 
opt.add_argument("--auto-open-devtools-for-tabs")
opt.add_experimental_option("excludeSwitches", ["enable-automation"])
opt.add_experimental_option('useAutomationExtension', False)
opt.add_argument("--window-size=1034,620")

#Seleciona o local de instalação do webdriver e as opções listadas acima.
driver = webdriver.Chrome(options=opt, executable_path= r'C:\Users\Take4\Documents\Saves\chromedriver_win32\chromedriver.exe')


#NOVOS CADASTROS - PUXADOS PELO DJANGO - NOVOS CADASTROS - PUXADOS PELO DJANGO - NOVOS CADASTROS - PUXADOS PELO DJANGO - 
#NOVOS CADASTROS - PUXADOS PELO DJANGO - NOVOS CADASTROS - PUXADOS PELO DJANGO - NOVOS CADASTROS - PUXADOS PELO DJANGO - 

#Aciono o link do Django, e autentico com nome de usuario e senha para o acesso.
driver.get("https://homolog.farmazon.com.br/")

mail_addressDjango = "-"
passwordDjango = "-"

#Autenticação no Django.
WebDriverWait(driver, 160).until(EC.element_to_be_clickable((By.NAME, "username"))).send_keys(mail_addressDjango)
WebDriverWait(driver, 160).until(EC.element_to_be_clickable((By.NAME, "password"))).send_keys(passwordDjango)
WebDriverWait(driver, 160).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/div/form[2]/button'))).click()

#Localizada a primeira pagina de novos usarios.
djangoInfo = driver.get("https://homolog.farmazon.com.br/api/customers/?format=json&limit=1000")

#Primeira pagina de novos usuarios é salva como JSON no arquivo TXT
with open('newUsers'+mes+ano+'.txt', 'w', encoding = 'utf-16') as f:
    f.write(driver.page_source)

arquivo1 = (r'C:\Users\Take4\Desktop\KPI\\newUsers'+mes+ano+'.txt')

#Arquivo TXT salvo é lido para ver quantas repeticoes do "create time" + ano + mes existem.
f1 = open(arquivo1, encoding = 'utf-16')
file_contents1 = f1.read()

countUsers = file_contents1.count('create_time":"' + ano + '-' + mes + '-')

totalNewUsers = countUsers

f1.close()

novosUsuariosInicio = r'C:\Users\Take4\Desktop\KPI\newUsers'+mes+ano+'.txt'
novosUsuariosFinal = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"
filename = os.path.basename(novosUsuariosInicio)
dest = os.path.join(novosUsuariosFinal,filename)
shutil.move(novosUsuariosInicio, dest)


#RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE
#RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE - RELATORIO GOOGLE

#RELATORIO INSTALACOES E DESINSTALACOES GOOGLE - RELATORIO INSTALACOES E DESINSTALACOES GOOGLE - RELATORIO INSTALACOES E DESINSTALACOES GOOGLE - 

#Transforma o arquivo baixado do Google (em CSV) em um arquivo XLSX.
#Arquvios do Google vem em UTF-16 - necessária a formatação específica.
path1Google = r'C:\Users\Take4\Downloads\stats_installs_installs_br.com.farmazon.client_'+ano+mes+'_overview.csv'
path2Google = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"+mes+ano+' GoogleInstalls&Uninstalls.xlsx'
transform = pd.read_csv(path1Google, sep="\t", encoding = 'utf-16')
transform.to_excel(path2Google, index=None)

#Formata o arquivos Google XLSX para que possamos usar um contador na coluna de instalaçoes e desinstalaçoes.
df = pd.read_excel(path2Google)
df['Date'], df['Package Name'], df['Daily Installs'], df['Daily Uninstalls'], df['Daily Device Upgrades'], df['Total User Installs'], df['Daily User Installs'], df['Daily User Uninstalls'], df['Active Device Installs'], df['0'], df['Update events'],df['00'] = df["Date,Package Name,Daily Device Installs,Daily Device Uninstalls,Daily Device Upgrades,Total User Installs,Daily User Installs,Daily User Uninstalls,Active Device Installs,Install events,Update events,Uninstall events"].str.split(',', 11).str
del df["Date,Package Name,Daily Device Installs,Daily Device Uninstalls,Daily Device Upgrades,Total User Installs,Daily User Installs,Daily User Uninstalls,Active Device Installs,Install events,Update events,Uninstall events"]


#Após formatado, arquivo é salvo com o mesmo nome (overwrite) 
df.to_excel(path2Google)

#Seleciono o arquivo e a planilha correta para contagem
google = load_workbook(path2Google)
googlews = google.get_sheet_by_name('Sheet1')  

#Calcula-se o total de instalações Google naquele mês (coluna K) - na variável "installsGoogle".
colunaInstalacoesGoogle = googlews['K']
listaNaColunaINT = [colunaInstalacoesGoogle[x].value for x in range(len(colunaInstalacoesGoogle))]
instalacoesGoogle = [int(x) for x in listaNaColunaINT]
installsGoogle = sum(instalacoesGoogle)

#Calcula-se o total de desinstalações Google naquele mês (coluna M) - na variável "uninstallsGoogle"
colunaDesinstalacoesGoogle = googlews['M']
listaNaColunaINT = [colunaDesinstalacoesGoogle[x].value for x in range(len(colunaDesinstalacoesGoogle))]
desinstalacoesGoogle = [int(x) for x in listaNaColunaINT]
uninstallsGoogle = sum(desinstalacoesGoogle)


#RELATORIO USUARIOS ATIVOS GOOGLE - RELATORIO USUARIOS ATIVOS GOOGLE - RELATORIO USUARIOS ATIVOS GOOGLE 

#Transforma o arquivo baixado do Google (em CSV) em um arquivo XLSX.
#Este arquivo do Google vem em UTF-8 - necessária a formatação específica.
path3Google = r'C:\Users\Take4\Downloads\Todos os países _ todas as regiões.csv'
path4Google = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"+mes+ano+' GoogleActiveUsers.xlsx'
transform = pd.read_csv(path3Google, sep="\t", encoding = 'utf-8')
transform.to_excel(path4Google, index=None)

#Remove o Header do arquivo.
replaceHeader = 'Data,"Público de instalação (Todos os usuários, Usuários únicos, Por intervalo, Diária): Todos os países / todas as regiões",Observações'


#Salva o arquivo na formatação correta.
df = pd.read_excel(path4Google)
df['Data'], df['0'], df['OBS'] = df['Data,"Público de instalação (Todos os usuários, Usuários únicos, Por intervalo, Diária): Todos os países / todas as regiões",Observações'].str.split(',', 2).str
del df[replaceHeader]

df.to_excel(path4Google)

#Seleciono o arquivo e a planilha correta para contagem
googleUsers = load_workbook(path4Google)
googlews = googleUsers.get_sheet_by_name('Sheet1') 

#Calcula-se o total de usuarios ativos (coluna C) Google naquele mês - na variável "UAGoogle"
colunaUsuariosAtivosGoogle = googlews['C']
listaNaColunaINTUAG = [colunaUsuariosAtivosGoogle[x].value for x in range(len(colunaUsuariosAtivosGoogle))]
usuariosAtivosGoogle1 = [str(x) for x in listaNaColunaINTUAG]

element = 0

while len(usuariosAtivosGoogle1) > element:
    if '.' in usuariosAtivosGoogle1[element]:
        numeroCorrigido = usuariosAtivosGoogle1[element].replace('.', '')
        del usuariosAtivosGoogle1[element]
        usuariosAtivosGoogle1.append(numeroCorrigido)
    else:
        pass
        element += 1



usuariosAtivosGoogle = [int(n) for n in usuariosAtivosGoogle1]


usersGoogle = sum(usuariosAtivosGoogle)
UAGoogle = usersGoogle/len(usuariosAtivosGoogle)



#RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE
#RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE - RELATORIO APPLE

#INSTALAÇOES APPLE - INSTALAÇOES APPLE - INSTALAÇOES APPLE - INSTALAÇOES APPLE - INSTALAÇOES APPLE:

#Transforma o arquivo baixado da Apple (em CSV) em um arquivo XLSX.
#Arquvios do Google vem em UTF-8 - necessária a formatação específica.

path1IApple = r'C:\Users\Take4\Downloads\farmazon___o_app_das_farmácias-instalações-'+ano+mes+diainicial+'-'+ano+mes+diafinal+'.csv'
path2IApple = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"+mes+ano+' AppleInstalls.xlsx'
transform = pd.read_csv(path1IApple, sep="\t", encoding = 'utf-8')
transform.to_excel(path2IApple, index=None)


#Formatcao arquivos XLSX após alteração de nome do app para v2
wbIApple = load_workbook(path2IApple)
wsIApple = wbIApple ['Sheet1']
wsIApple.cell(row=1, column=1).value = "Nome,Farmazon"
wbIApple.save(path2IApple)

#Formata o arquivos Google XLSX para que possamos usar um contador na coluna de instalaçoes.
df = pd.read_excel(path2IApple)
df['Date'], df['0'] = df["Nome,Farmazon"].str.split(',', 1).str
del df["Nome,Farmazon"]


#Após formatado, arquivo é salvo com o mesmo nome (overwrite) 
df.to_excel(path2IApple)

#Necessária formatação extra, pois as primeiras linhas vem com informações desnecessárias.
appleI = pd.read_excel(path2IApple, sep='|', skiprows=list(range(1,4)))
appleI.to_excel(path2IApple)

#Seleciono o arquivo e a planilha correta para contagem
appleI = load_workbook(path2IApple)
appleIws = appleI.get_sheet_by_name('Sheet1')  

#Calcula-se o total de desinstalações (coluna D) Apple naquele mês - na variável "installsApple"
colunaInstalacoesApple = appleIws['D']
listaNaColunaINT = [colunaInstalacoesApple[x].value for x in range(len(colunaInstalacoesApple))]
instalacoesApple = [int(x) for x in listaNaColunaINT]
installsApple = sum(instalacoesApple)

#DESINSTALAÇOES APPLE - DESINSTALAÇOES APPLE - DESINSTALAÇOES APPLE - DESINSTALAÇOES APPLE - DESINSTALAÇOES APPLE 

#Transforma o arquivo baixado da Apple (em CSV) em um arquivo XLSX.
#Arquvios do Google vem em UTF-18 - necessária a formatação específica.
path1UApple = r'C:\Users\Take4\Downloads\farmazon___o_app_das_farmácias-desinstalações-'+ano+mes+diainicial+'-'+ano+mes+diafinal+'.csv'
path2UApple = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"+mes+ano+' AppleUninstalls.xlsx'
transform = pd.read_csv(path1UApple, sep="\t", encoding = 'utf-8')
transform.to_excel(path2UApple, index=None)

#Formatcao arquivos XLSX após alteração de nome do app para v2
wbUApple = load_workbook(path2UApple)
wsUApple = wbUApple ['Sheet1']
wsUApple.cell(row=1, column=1).value = "Nome,Farmazon"
wbUApple.save(path2UApple)


#Formata o arquivos Google XLSX para que possamos usar um contador na coluna de instalaçoes.
df = pd.read_excel(path2UApple)
df['Date'], df['0'] = df["Nome,Farmazon"].str.split(',', 1).str
del df["Nome,Farmazon"]


#Após formatado, arquivo é salvo com o mesmo nome (overwrite) 
df.to_excel(path2UApple)

#Necessária formatação extra, pois as primeiras linhas vem com informações desnecessárias.
appleU = pd.read_excel(path2UApple, sep='|', skiprows=list(range(1,4)))
appleU.to_excel(path2UApple)

#Seleciono o arquivo e a planilha correta para contagem
appleD = load_workbook(path2UApple)
appleDws = appleD.get_sheet_by_name('Sheet1')  

#Calcula-se o total de desinstalações (coluna D) Apple naquele mês - na variável "uninstallsApple"
colunaDesinstalacoesApple = appleDws['D']
listaNaColunaINT = [colunaDesinstalacoesApple[x].value for x in range(len(colunaDesinstalacoesApple))]
desinstalacoesApple = [int(x) for x in listaNaColunaINT]
uninstallsApple = sum(desinstalacoesApple)


#USUARIOS ATIVOS APPLE - USUARIOS ATIVOS APPLE - USUARIOS ATIVOS APPLE - USUARIOS ATIVOS APPLE - USUARIOS ATIVOS APPLE

#Transforma o arquivo baixado da Apple (em CSV) em um arquivo XLSX.
#Este arquivo da Apple vem em UTF-8 - necessária a formatação específica.
path1UAApple = r'C:\Users\Take4\Downloads\farmazon___o_app_das_farmácias-dispositivos_ativos-'+ano+mes+diainicial+'-'+ano+mes+diafinal+'.csv'
path2UAApple = r'C:\Users\Take4\Desktop\KPI\\'+mes+ano+"\\"+mes+ano+' AppleActiveUsers.xlsx'
transform = pd.read_csv(path1UAApple, sep="\t", encoding = 'utf-8')
transform.to_excel(path2UAApple, index=None)

#Formatcao arquivos XLSX após alteração de nome do app para v2
wbUAApple = load_workbook(path2UAApple)
wsUAApple = wbUAApple ['Sheet1']
wsUAApple.cell(row=1, column=1).value = "Nome,Farmazon"
wbUAApple.save(path2UAApple)

#Formata o arquivos Apple para XLSX para que possamos usar um contador na coluna de instalaçoes.
df = pd.read_excel(path2UAApple)
df['Date'], df['0'] = df["Nome,Farmazon"].str.split(',', 1).str
del df["Nome,Farmazon"]

#Após formatado, arquivo é salvo com o mesmo nome (overwrite) 
df.to_excel(path2UAApple)

#Necessária formatação extra, pois as primeiras linhas vem com informações desnecessárias.
appleUA = pd.read_excel(path2UAApple, sep='|', skiprows=list(range(1,4)))
appleUA.to_excel(path2UAApple)

#Seleciono o arquivo e a planilha correta para contagem
appleUA = load_workbook(path2UAApple)
appleUAws = appleUA.get_sheet_by_name('Sheet1')  

#Calcula-se o total de usuarios ativos (coluna D) Apple naquele mês - na variável "UAApple"
colunaUAApple = appleUAws['D']
listaNaColunaINT = [colunaUAApple[x].value for x in range(len(colunaUAApple))]
usuariosAtivosApple = [int(x) for x in listaNaColunaINT]
UAApple = sum(usuariosAtivosApple)


#TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL
#TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL - TOTAL

#Total de Novos Cadastros puxados no Django
totalNovosCadastros = totalNewUsers

#Total de instalações Google + Apple - totalInstalls.
totalInstalls = installsApple + installsGoogle
print ('Instalacoes Apple')
print (installsApple)
print (installsGoogle)

#Total de desinstalações Google + Apple - totalUninstalls.
totalUninstalls = uninstallsApple + uninstallsGoogle
print ('Unistalls Apple')
print (uninstallsApple)
print (uninstallsGoogle)

#Total Usuarios Ativos - UAGoogle + UAApple + devida formatação por conta de casas decimais.
totalUsers1 = UAGoogle + UAApple
totalUsers = ("%.0f" % totalUsers1)
print (UAGoogle)
print ('UA Apple')
print (UAApple)

#Linha correspondentes no KPI.
linhaInstalls = 3
linhaActiveUsers = 4
linhaNovosUsuarios = 5
linhaUninstalls = 6

#Coluna do mês correspondente.
if ano == "2019":
    Mes = int(mes) - 5
if ano == "2020":
    Mes = int(mes) + 7
if ano == "2021":
    Mes = int(mes) + 19
elif ano == "2022":
    Mes = int(mes) + 31

#Atualiza na tabela do Google Sheet os valores.
sheet.update_cell (linhaInstalls, Mes, totalInstalls)
sheet.update_cell (linhaActiveUsers, Mes, totalUsers)
sheet.update_cell (linhaNovosUsuarios, Mes, totalNovosCadastros)
sheet.update_cell (linhaUninstalls, Mes, totalUninstalls)

driver.close()

print (totalUninstalls)



